"""Parse one employee's weekly timesheet workbook.

The Excel template (`UC_SRE_Timesheet_Templet_2026.xlsx`) holds exactly ONE
week per file. Layout reverse-engineered from the live template:

    row  4 col F : Week Starting (datetime, Monday)
    row  4 col C : Employee Name (informational)
    row  5 col C : Employee ID
    row 12       : header row ("Main Category | Sub-Category | Project No. | Mon..Sun | Task Description | Total")
    row 13       : date axis ("Week of:" + per-day dates)
    rows 14..28  : entry rows (up to 15 entries)
    row 29       : "Total Hours Worked" footer

History import = one file per (employee, week). The CLI iterates a folder of
files and calls this parser per file.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..schema import MainCategory, WeekEntryRow, WeekRow


class WorkbookParseError(ValueError):
    """Raised when the workbook is not a recognizable SRE timesheet."""


SHEET_NAME = "Weekly Timesheet"
HEADER_ROW = 12
DATE_ROW = 13
FIRST_ENTRY_ROW = 14
LAST_ENTRY_ROW = 28

# Column indices (1-based, matches openpyxl)
COL_MAIN = 2     # B
COL_SUB = 3      # C
COL_PROJ = 4     # D
COL_MON = 5      # E
COL_SUN = 11     # K
COL_DESC = 12    # L

# Cell coordinates for header values
CELL_WEEK_START = ("F", 4)
CELL_EMP_ID = ("C", 5)


VALID_MAIN: set[MainCategory] = {"Project", "Admin", "Office & Sales"}


def parse_workbook(path: Path) -> tuple[str, WeekRow]:
    """Parse one workbook → (employee_code, WeekRow).

    `data_only=True` resolves formula cells to their last cached value, so
    `Total hrs` and other computed cells render as numbers instead of formulas.
    """
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
    except Exception as e:  # openpyxl raises a grab-bag of exceptions
        raise WorkbookParseError(f"could not open {path.name}: {e}") from e

    if SHEET_NAME not in wb.sheetnames:
        raise WorkbookParseError(
            f"{path.name}: missing sheet {SHEET_NAME!r} "
            f"(found: {wb.sheetnames})"
        )
    ws = wb[SHEET_NAME]

    week_start = _read_week_start(ws, path.name)
    employee_code = _read_employee_code(ws, path.name)
    entries = _read_entries(ws, path.name)

    return employee_code, WeekRow(week_start=week_start, entries=entries)


def _read_week_start(ws: Any, fname: str) -> date:
    col, row = CELL_WEEK_START
    raw = ws[f"{col}{row}"].value
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return date.fromisoformat(raw.strip())
        except ValueError as e:
            raise WorkbookParseError(
                f"{fname}: week_start cell {col}{row} is not a date: {raw!r}"
            ) from e
    raise WorkbookParseError(
        f"{fname}: week_start cell {col}{row} is empty or not a date (got {type(raw).__name__})"
    )


def _read_employee_code(ws: Any, fname: str) -> str:
    col, row = CELL_EMP_ID
    raw = ws[f"{col}{row}"].value
    if raw is None or str(raw).strip() == "":
        raise WorkbookParseError(f"{fname}: missing employee id at {col}{row}")
    return str(raw).strip()


def _read_entries(ws: Any, fname: str) -> list[WeekEntryRow]:
    entries: list[WeekEntryRow] = []
    for r in range(FIRST_ENTRY_ROW, LAST_ENTRY_ROW + 1):
        main_raw = ws.cell(r, COL_MAIN).value
        if main_raw is None:
            continue
        main = str(main_raw).strip()
        if not main:
            continue

        if main not in VALID_MAIN:
            raise WorkbookParseError(
                f"{fname} row {r}: unknown main category {main!r} "
                f"(expected one of {sorted(VALID_MAIN)})"
            )

        sub_raw = ws.cell(r, COL_SUB).value
        sub = str(sub_raw).strip() if sub_raw is not None else ""
        if not sub:
            raise WorkbookParseError(
                f"{fname} row {r}: main category {main!r} requires a sub-category"
            )

        hours = [_to_decimal(ws.cell(r, c).value) for c in range(COL_MON, COL_SUN + 1)]
        desc_raw = ws.cell(r, COL_DESC).value
        desc = str(desc_raw).strip() if desc_raw is not None else ""

        total = sum(hours)
        if total == 0 and not desc:
            # Genuinely empty row inside the entry block — skip.
            continue
        if total > 0 and not desc:
            raise WorkbookParseError(
                f"{fname} row {r}: hours logged ({total}) but description is empty"
            )

        project_raw = ws.cell(r, COL_PROJ).value
        project_number: int | None = None
        if project_raw not in (None, ""):
            try:
                project_number = int(project_raw)
            except (TypeError, ValueError) as e:
                raise WorkbookParseError(
                    f"{fname} row {r}: project number must be an integer, got {project_raw!r}"
                ) from e

        if main == "Project" and project_number is None:
            raise WorkbookParseError(
                f"{fname} row {r}: main category 'Project' requires a project number"
            )

        entries.append(
            WeekEntryRow(
                main_category=main,  # type: ignore[arg-type]  # narrowed by VALID_MAIN check
                sub_category_name=sub,
                project_number=project_number,
                mon_hrs=hours[0],
                tue_hrs=hours[1],
                wed_hrs=hours[2],
                thu_hrs=hours[3],
                fri_hrs=hours[4],
                sat_hrs=hours[5],
                sun_hrs=hours[6],
                description=desc,
                position=len(entries),
            )
        )

    if not entries:
        raise WorkbookParseError(f"{fname}: no entry rows found in {SHEET_NAME!r}")
    return entries


def _to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:  # noqa: BLE001 — convert any malformed input to 0
        return Decimal("0")
