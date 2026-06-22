from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from sre_import.parse.workbook import parse_workbook, WorkbookParseError

FIXTURE = Path(__file__).parent / "fixtures" / "employee_sample.xlsx"


def test_parses_real_template():
    code, week = parse_workbook(FIXTURE)
    assert code == "1"
    assert week.week_start == date(2026, 3, 30)
    assert week.week_start.weekday() == 0  # Monday
    assert len(week.entries) == 6  # 6 logged rows in the live template


def test_known_entry_values():
    _, week = parse_workbook(FIXTURE)
    by_desc = {e.description: e for e in week.entries}
    travel = by_desc["Travel to YYC"]
    assert travel.main_category == "Project"
    assert travel.sub_category_name == "Site Travel"
    assert travel.project_number == 2026125
    assert travel.mon_hrs == Decimal("8")
    assert travel.sat_hrs == Decimal("0")


def test_project_main_without_project_number_rejected(tmp_path):
    wb = _scaffold()
    ws = wb["Weekly Timesheet"]
    ws.cell(14, 2).value = "Project"
    ws.cell(14, 3).value = "Site Work"
    ws.cell(14, 5).value = 4
    ws.cell(14, 12).value = "Description"
    bad = tmp_path / "no_project.xlsx"
    wb.save(bad)
    with pytest.raises(WorkbookParseError, match="requires a project number"):
        parse_workbook(bad)


def test_hours_without_description_rejected(tmp_path):
    wb = _scaffold()
    ws = wb["Weekly Timesheet"]
    ws.cell(14, 2).value = "Admin"
    ws.cell(14, 3).value = "Sick Time"
    ws.cell(14, 5).value = 8
    bad = tmp_path / "no_desc.xlsx"
    wb.save(bad)
    with pytest.raises(WorkbookParseError, match="description is empty"):
        parse_workbook(bad)


def test_unknown_main_category_rejected(tmp_path):
    wb = _scaffold()
    ws = wb["Weekly Timesheet"]
    ws.cell(14, 2).value = "Travel"  # Not in {Project, Admin, Office & Sales}
    ws.cell(14, 3).value = "x"
    ws.cell(14, 12).value = "x"
    ws.cell(14, 5).value = 1
    bad = tmp_path / "bad_main.xlsx"
    wb.save(bad)
    with pytest.raises(WorkbookParseError, match="unknown main category"):
        parse_workbook(bad)


def test_missing_sheet_rejected(tmp_path):
    wb = Workbook()
    wb.active.title = "Random"
    bad = tmp_path / "no_sheet.xlsx"
    wb.save(bad)
    with pytest.raises(WorkbookParseError, match="missing sheet"):
        parse_workbook(bad)


def test_empty_entry_block_rejected(tmp_path):
    wb = _scaffold()
    bad = tmp_path / "empty.xlsx"
    wb.save(bad)
    with pytest.raises(WorkbookParseError, match="no entry rows"):
        parse_workbook(bad)


def _scaffold() -> Workbook:
    """Build a minimal workbook matching the parser's expected layout."""
    wb = Workbook()
    wb.active.title = "Weekly Timesheet"
    ws = wb["Weekly Timesheet"]
    ws["F4"] = date(2026, 1, 5)
    ws["C5"] = "E001"
    return wb
